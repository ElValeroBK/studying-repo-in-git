import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def result_workbook(file_book):
     ws = xl.load_workbook(file_book)
     sheet = ws['Sheet1']
     #cell = sheet['a1']
     cell = sheet.cell(1,2)

     for row in range(2,sheet.max_row + 1):
          cell = sheet.cell(row,3)
          value_cell = cell.value * 0.9
          value_real = sheet.cell(row,4)
          value_real.value = value_cell

     value = Reference(sheet,
               min_row=2,
               max_row=sheet.max_row,
               min_col=4,
               max_col=4)

     chart = BarChart()
     chart.add_data(value)
     sheet.add_chart(chart,'e2')

     ws.save("transaction2.xlsx")


result_workbook('transactions.xlsx')
