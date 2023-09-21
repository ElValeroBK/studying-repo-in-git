# ### File Handling ###

import csv
import json
import os

# .txt file

# Leer, escribir y sobrescribir si ya existe
txt_file = open("my_file.txt", "w+")

txt_file.write("\nName: Yasiel\nApellido: Valero\nAge: 37 year old\nY mi lenguaje preferido es Python\n")
txt_file.write("Name\n")

print(txt_file.read())   #read all
print(txt_file.read(10)) #read and stay in position
print(txt_file.readline()) # read line por line
print(txt_file.readline())
for line in txt_file.readlines():
   print(line)

txt_file.write("\nAunque también me gusta Kotlin")
print(txt_file.readline())

txt_file.close()


with open("my_file.txt", "a") as yasie:
    yasie.write("\nHello I am yasi")

# os.remove("my_file.txt")



# Clase en vídeo (03/11/22): https://www.twitch.tv/videos/1642512950

# .json file


json_file = open("my_file.json", "w+")

json_test = {
    "name": "Brais",
    "surname": "Moure",
    "age": 35,
    "languages": ["Python", "Swift", "Kotlin"],
    "website": "https://moure.dev"}

json.dump(json_test, json_file, indent=2 )

json_file.close()

with open("my_file.json") as my_other_file:
    for line in my_other_file.readlines():
        print(line)

json_dict = json.load(open("my_file.json"))
print(json_dict)
print(type(json_dict))
print(json_dict["name"])



# # .csv file
csv_file = open("my_file.csv", "w+")

csv_writer = csv.writer(csv_file)
csv_writer.writerow(["name", "surname", "age", "language", "website"])
csv_writer.writerow(["Brais", "Moure", 35, "Python", "https://moure.dev"])
csv_writer.writerow(["Roswell", "", 2, "COBOL", ""])

csv_file.close()

with open("my_file.csv") as my_other_file:
    for line in my_other_file.readlines():
        print(line)

# .xlsx file
# import xlrd # Debe instalarse el módulo

import xlsxwriter


workbook = xlsxwriter.Workbook('my_file.xlsx')
worksheet = workbook.add_worksheet("Yasielsheet")
worksheet2 = workbook.add_worksheet("helloworld")
 
# Start from the first cell.
# Rows and columns are zero indexed.
row = 0
column = 0
 
content = (
    ['ankit', 1000],
    ['rahul',   100],
    ['priya',  300],
    ['harshita',    50],
)
 
# iterating through content list
for name, score in content :
 
    # write operation perform
    worksheet.write(row, column, name)
    worksheet.write(row, column + 1, score)
 
    # incrementing the value of row by one
    # with each iterations.
    row += 1

workbook.close()

import openpyxl
from tabulate import tabulate

data= []

exelopen = openpyxl.load_workbook("my_file.xlsx")
worksheetnew = exelopen.active

for row in range(0, worksheetnew.max_row):
    _row = [row,]
    
    for col in worksheetnew.iter_cols(0,worksheetnew.max_column):
        _row.append(col[row].value)

    data.append(_row)
headers = ["id", "name", "value","color"]
header_align = (("center",) * worksheetnew.max_column)

print(tabulate(data,headers=headers,tablefmt= 'heavy_grid',colalign= header_align ))

exelopen.close()

# .xml file

# ¿Te atreves a practicar cómo trabajar con este tipo de ficheros?
import xml.etree.cElementTree as ET

root = ET.Element("my_file.xml")
doc = ET.SubElement(root, "doc")

ET.SubElement(doc, "field1", name="blah").text = "some value1"
ET.SubElement(doc, "field2", name="asdfasd").text = "some vlaue2"

tree = ET.ElementTree(root)
tree.write("my_file.xml")
